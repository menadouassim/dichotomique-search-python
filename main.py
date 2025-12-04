import random
import time

def dechatomique_search(list, target):
    
    start=0
    end=len(list)-1
    
    while start<=end :
        mid=(start+end)//2
        if list[mid]==target:
            
            return mid
        elif list[mid]<target:
            start=mid+1
            
        else:
            end=mid-1
            
    return "target not found"
def linear_search(list, target):
    for i in range(len(list)):
        if list[i]==target:
            return i
    return "target not found"
def generate_sorted_list(size):
    return sorted(random.sample(range(1, size + 1), size))
# Example usage:


def get_time(n):
    my_list = generate_sorted_list(n)
    target_value = random.randint(1, 1_000_000)
    time_start = time.time()
    dechatomique_search_result = dechatomique_search(list=my_list, target=target_value)
    time_end = time.time()
    dechatomique_time = time_end - time_start
    # Measure linear search
    time_start = time.time()
    linear_search_result = linear_search(list=my_list, target=target_value)
    time_end = time.time()
    linear_time = time_end - time_start 
    return dechatomique_time, linear_time

from openpyxl import Workbook


list_n = [10, 1_000,10_000,100_000, 1_000_000,10_000_000]

# Create a workbook and sheet
wb = Workbook()
ws = wb.active
ws.title = "Search Comparison"

# Add headers


ws.cell(row=2, column=1, value="dechatomique Search Time (seconds)")
ws.cell(row=3, column=1, value="Linear Search Time (seconds)")
# Add data
for n in list_n:
    dechatomique_time,linear_time=get_time(n)
    next_row = ws.max_column + 1
    ws.cell(row=1, column=next_row, value=f"n={n}")
    ws.cell(row=2, column=next_row, value=dechatomique_time) 
    ws.cell(row=3, column=next_row, value=linear_time)
    

# Save Excel file
wb.save("search_comparison.xlsx")

print("Results exported to search_comparison.xlsx")
